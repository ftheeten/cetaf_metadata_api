from django.core.management.base import BaseCommand, CommandError
from django.conf import settings
from  urllib.parse import quote
import requests
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from io import BytesIO
#pip install odfpy
import pandas as pnd
import numpy as np
import openpyxl
import sys
import traceback
import urllib

import pygsheets
import gdown
from collections import OrderedDict



class read_excel():
    def __init__(self):
        self.refresh_token=settings.GOOGLE_CLOUD_REFRESH_TOKEN
        self.client_id     = settings.GOOGLE_CLOUD_CLIENT_ID
        self.client_secret = settings.GOOGLE_CLOUD_SECRET
        self.redirect_uri=settings.GOOGLE_CLOUD_REDIRECT_URI
        self.gs_auth_file=settings.GOOGLE_AUTH_FILE

        self.refresh_url = "https://www.googleapis.com/oauth2/v4/token"
        self.encoded_client=quote(self.client_id, safe='~()*!.\'')
        self.encoded_secret=quote(self.client_secret, safe='~()*!.\'')
        self.encoded_token=quote(self.refresh_token, safe='~()*!.\'')
        
    def get_excel(self, file_id, sheet_name, p_header=None, skiprows=0, engine=None):
        fh=self.download_logic( file_id)
        if fh is not None:
            if engine is None:
                df = pnd.read_excel(fh.getvalue(), sheet_name=sheet_name, header=p_header, skiprows=skiprows)
            else:
                df = pnd.read_excel(fh.getvalue(), sheet_name=sheet_name, header=p_header, skiprows=skiprows, engine=engine)
            df = df.replace({np.nan: None})
            wb = openpyxl.load_workbook(fh)
            print(wb.properties.modified)
            return df, wb.properties.modified
        return None, None        
        
 
    def get_gs_by_id(self, file_id, sheet_name=None, p_header=None, skiprows=0, engine=None):
        fh=self.download_logic( file_id)
        df = pnd.read_excel( BytesIO(fh.getvalue()), sheet_name=sheet_name, header=p_header, skiprows=skiprows)
        return df   
 
    def get_excel_sheet_names(self, file_id):        
        returned=None
        fh=self.download_logic( file_id)
        if fh:
            wb = openpyxl.load_workbook(fh)
            returned= wb.sheetnames  
        return returned
        
    def download_logic(self, file_id):
        returned = None
        headers={'Content-Type': 'Content-Type: application/x-www-form-urlencoded',}
        post_body = 'grant_type=refresh_token&client_id='+self.encoded_client+'&client_secret='+self.encoded_secret+'&refresh_token='+self.encoded_token
        resp=requests.post(self.refresh_url, params=post_body, headers=headers)
        if resp.status_code==200:
            access_token=resp.json()["access_token"]
            scopes=["https://www.googleapis.com/auth/drive"]
            creds = Credentials(token=access_token, default_scopes =scopes, client_id =self.client_id, client_secret =self.client_secret  )
            if not creds or not creds.valid:
                print("invalid")
            else:
                try: 
                    service = build('drive', 'v3', credentials=creds)
                    request = service.files().get_media(fileId=file_id)#.execute()
                    fh = BytesIO()
                    downloader = MediaIoBaseDownload(fd=fh, request=request)
                    done = False
                    while not done:
                        status, done = downloader.next_chunk()
                    returned=fh
                except HttpError as error:
                    # TODO(developer) - Handle errors from drive API.
                    print(f'An error occurred: {error}')
                    traceback.print_exc()
                    
        return returned
        

        
    def get_gs_xls_from_url(self, p_url, sheet_name=None):
        returned=None
        try:
            byio = BytesIO()
            gdown.download(p_url, byio, quiet=False)
            #print(byio.getvalue())
            #tmp =byio
            
            returned = pnd.read_excel(byio, engine='openpyxl', sheet_name=sheet_name )
            byio.close()
        except HttpError as error:
            # TODO(developer) - Handle errors from drive API.
            print(f'An error occurred: {error}')
            traceback.print_exc()
            print("EXIT_EXCEPTION XLS")
            #sys.exit()   
        return returned
        
        
    def get_gs_ods_from_url(self, p_url, sheet_name=None):
        returned=None
        try:
            byio = BytesIO()
            gdown.download(p_url, byio, quiet=False)
            print(byio.getvalue())
            #tmp =byio
            #returned = pnd.read_excel(byio, engine="pyxlsb", sheet_name=sheet_name )
            returned = pnd.read_excel(byio, engine="odf", sheet_name=sheet_name )
            byio.close()
        except HttpError as error:
            # TODO(developer) - Handle errors from drive API.
            print(f'An error occurred: {error}')
            traceback.print_exc()
            print("EXIT_EXCEPTION")
            #sys.exit()   
        return returned
        
     
    #standard google sheet
    def read_gsheets(self, file_id):
        returned= OrderedDict()
        #print(self.gs_auth_file)
        client = pygsheets.authorize(service_file=self.gs_auth_file)
        sh = client.open_by_key(file_id)
        wks=sh.worksheets()
        for w in wks:
            tmp=w.get_as_df()
            returned[w.title]=tmp        
        return returned
  
    def panda_unique_cols(self, p_pnd):
        renamer = {}
        column_names = list(p_pnd.columns)
        for col in column_names:
            col=str(col)
            if col.strip()=="":
                col="unnamed"
            if not col in renamer:
                renamer[col]=[1]
            else:
                renamer[col].append(renamer[col][-1]+1)
        new_cols=[]
        for col in column_names: 
            col=str(col)
            if col.strip()=="":
                col="unnamed"
            cpt= renamer[col].pop(0)
            if cpt>1:
                col=col+'_'+str(cpt)
            new_cols.append(col)
        
        p_pnd.columns=new_cols
        return p_pnd
        